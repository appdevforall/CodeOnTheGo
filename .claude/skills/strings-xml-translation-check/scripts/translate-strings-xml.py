#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "google-genai",
#   "google-cloud-translate",
#   "openpyxl",
# ]
# ///
"""Translate Android strings.xml foreign values back to English for QA review.

Reads an English strings.xml and a foreign-language strings.xml, round-trips
each foreign value back to English via Gemini and Google Cloud Translation,
and writes a side-by-side .xlsx for visual inspection.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

from google import genai
from google.cloud import translate_v2 as translate
from google.genai import types
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


HEADER = ["key", "english", "foreign", "gemini_en", "google_en"]


def _element_text(el: ET.Element) -> str:
    """Return all text inside an element, flattening any nested markup."""
    return "".join(el.itertext())


def parse_strings_xml(path: Path) -> dict[str, str]:
    """Parse an Android strings.xml. Returns {key: value}.

    - <string name="X">value</string>  -> {"X": value}
    - <string-array name="X"><item>a</item><item>b</item></string-array>
      -> {"X[0]": "a", "X[1]": "b"}
    - Skips elements with translatable="false".
    - Skips <plurals> and any other element types.
    """
    tree = ET.parse(path)
    root = tree.getroot()
    out: dict[str, str] = {}

    for el in root:
        if el.get("translatable") == "false":
            continue
        name = el.get("name")
        if not name:
            continue
        if el.tag == "string":
            out[name] = _element_text(el)
        elif el.tag == "string-array":
            for i, item in enumerate(el.findall("item")):
                out[f"{name}[{i}]"] = _element_text(item)

    return out


def detect_source_lang(foreign_path: Path) -> str:
    """Derive a BCP-47-ish code from a values-XX or values-XX-rYY directory."""
    parent = foreign_path.parent.name
    if not parent.startswith("values-"):
        raise ValueError(
            f"cannot infer language from {parent!r}; pass --source-lang explicitly"
        )
    code = parent[len("values-"):]
    return re.sub(r"-r([A-Z]{2})$", r"-\1", code)


def gemini_translate_batch(
    client: genai.Client, model: str, source_lang: str, texts: list[str]
) -> list[str]:
    """Translate a batch via Gemini, returning one English string per input.

    Uses JSON output for reliable order preservation. On batch failure, falls
    back to per-string calls so a single bad input doesn't tank the run.
    """
    if not texts:
        return []

    prompt = (
        f"You are translating Android UI strings from {source_lang} to English "
        "for quality-assurance review.\n\n"
        "Rules:\n"
        "- Preserve format placeholders exactly: %s, %d, %1$s, %2$d, etc.\n"
        "- Preserve backslash escape sequences as literal text: \\n, \\', \\\", \\u2022.\n"
        "- Preserve any HTML/XML tags unchanged.\n"
        "- Output ONLY a JSON array of strings, one translation per input, "
        "in the same order as the input array. No commentary, no keys.\n\n"
        "Inputs (JSON array):\n"
        f"{json.dumps(texts, ensure_ascii=False)}"
    )

    try:
        resp = client.models.generate_content(
            model=model,
            contents=prompt,
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                temperature=0,
            ),
        )
        data = json.loads(resp.text)
        if not isinstance(data, list) or len(data) != len(texts):
            raise ValueError(
                f"expected list of {len(texts)}, got "
                f"{type(data).__name__} "
                f"len={len(data) if isinstance(data, list) else 'n/a'}"
            )
        return [str(x) for x in data]
    except Exception as e:
        print(f"  [gemini batch failed: {e}] falling back per-string",
              file=sys.stderr)
        return [_gemini_translate_one(client, model, source_lang, t) for t in texts]


def _gemini_translate_one(
    client: genai.Client, model: str, source_lang: str, text: str
) -> str:
    if not text:
        return ""
    prompt = (
        f"Translate this Android UI string from {source_lang} to English. "
        "Preserve format placeholders (%1$s, %d, %s) and backslash escapes "
        "(\\n, \\u2022) exactly. Reply with only the translation, no commentary.\n\n"
        f"Input: {text}"
    )
    try:
        resp = client.models.generate_content(
            model=model,
            contents=prompt,
            config=types.GenerateContentConfig(temperature=0),
        )
        return (resp.text or "").strip()
    except Exception as e:
        print(f"    [gemini single failed for {text!r}: {e}]", file=sys.stderr)
        return ""


def google_translate_batch(
    client: translate.Client, source_lang: str, texts: list[str]
) -> list[str]:
    if not texts:
        return []
    try:
        results = client.translate(
            texts,
            target_language="en",
            source_language=source_lang,
            format_="text",
        )
        if isinstance(results, dict):
            results = [results]
        return [r.get("translatedText", "") for r in results]
    except Exception as e:
        print(f"  [google translate batch failed: {e}]", file=sys.stderr)
        return [""] * len(texts)


def write_xlsx(rows: list[tuple[str, str, str, str, str]], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"

    ws.append(HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append(row)

    for col_idx, header in enumerate(HEADER, start=1):
        max_len = len(header)
        for row in rows:
            v = row[col_idx - 1]
            if v and len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

    wb.save(output)


def main() -> int:
    p = argparse.ArgumentParser(
        description=(
            "Translate Android strings.xml foreign values back to English via "
            "Gemini and Google Translate; emit a side-by-side xlsx for QA review."
        )
    )
    p.add_argument("english_xml", type=Path,
                   help="path to English values/strings.xml")
    p.add_argument("foreign_xml", type=Path,
                   help="path to foreign values-XX/strings.xml")
    p.add_argument("--output", type=Path, default=None,
                   help="output xlsx path (default: ./translations-<lang>.xlsx)")
    p.add_argument("--source-lang", default=None,
                   help="BCP-47 language code (default: auto-detect from foreign path)")
    p.add_argument("--batch-size", type=int, default=50,
                   help="strings per Gemini batch (default: 50)")
    p.add_argument("--model", default="gemini-2.5-flash",
                   help="Gemini model (default: gemini-2.5-flash)")
    args = p.parse_args()

    if not args.english_xml.is_file():
        print(f"error: not a file: {args.english_xml}", file=sys.stderr)
        return 2
    if not args.foreign_xml.is_file():
        print(f"error: not a file: {args.foreign_xml}", file=sys.stderr)
        return 2

    gemini_key = os.environ.get("GEMINI_API_KEY")
    if not gemini_key:
        print("error: GEMINI_API_KEY env var is required", file=sys.stderr)
        return 2
    # Google Cloud Translation uses Application Default Credentials. Run
    # `gcloud auth application-default login` once, then
    # `gcloud auth application-default set-quota-project YOUR_PROJECT` so the
    # client has a quota project. GOOGLE_APPLICATION_CREDENTIALS (service
    # account JSON) is also supported if set, but not required.

    source_lang = args.source_lang or detect_source_lang(args.foreign_xml)
    output = args.output or Path.cwd() / f"translations-{source_lang}.xlsx"

    print(f"english:     {args.english_xml}")
    print(f"foreign:     {args.foreign_xml}")
    print(f"source lang: {source_lang}")
    print(f"output:      {output}")
    print(f"model:       {args.model}")

    english = parse_strings_xml(args.english_xml)
    foreign = parse_strings_xml(args.foreign_xml)
    print(f"english keys: {len(english)}, foreign keys: {len(foreign)}")

    all_keys = sorted(set(english) | set(foreign))
    keys_to_translate = [k for k in all_keys if foreign.get(k)]
    texts_to_translate = [foreign[k] for k in keys_to_translate]
    print(f"total rows: {len(all_keys)}, to translate: {len(keys_to_translate)}")

    gemini_client = genai.Client(api_key=gemini_key)
    gemini_results: dict[str, str] = {}
    bsize = args.batch_size
    n_batches = (len(texts_to_translate) + bsize - 1) // bsize
    for i in range(0, len(texts_to_translate), bsize):
        batch_keys = keys_to_translate[i:i + bsize]
        batch_texts = texts_to_translate[i:i + bsize]
        print(f"  gemini batch {i // bsize + 1}/{n_batches}: "
              f"{len(batch_texts)} strings...", flush=True)
        results = gemini_translate_batch(
            gemini_client, args.model, source_lang, batch_texts
        )
        for k, v in zip(batch_keys, results):
            gemini_results[k] = v

    g_client = translate.Client()
    google_results: dict[str, str] = {}
    gbsize = 100
    g_batches = (len(texts_to_translate) + gbsize - 1) // gbsize
    for i in range(0, len(texts_to_translate), gbsize):
        batch_keys = keys_to_translate[i:i + gbsize]
        batch_texts = texts_to_translate[i:i + gbsize]
        print(f"  google batch {i // gbsize + 1}/{g_batches}: "
              f"{len(batch_texts)} strings...", flush=True)
        results = google_translate_batch(g_client, source_lang, batch_texts)
        for k, v in zip(batch_keys, results):
            google_results[k] = v

    rows = [
        (
            k,
            english.get(k, ""),
            foreign.get(k, ""),
            gemini_results.get(k, ""),
            google_results.get(k, ""),
        )
        for k in all_keys
    ]

    write_xlsx(rows, output)
    print(f"\nwrote {len(rows)} rows -> {output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
